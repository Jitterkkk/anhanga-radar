import json
import openpyxl
from groq import Groq
from excel_manager import adicionar_linha, garantir_excel


def _aba_para_status(status: str) -> str:
    s = status.lower()
    if any(k in s for k in ["interesse", "recusou"]):
        return "Sem Interesse"
    return "Falhas e Sem Contato"


def _mapear_colunas(cabecalhos, amostra, api_key):
    client = Groq(api_key=api_key)

    prompt = f"""Você receberá cabeçalhos e linhas de exemplo de uma planilha Excel com dados de vereadores.
Mapeie cada campo destino ao nome EXATO do cabeçalho correspondente. Retorne SOMENTE JSON válido, sem markdown.

Campos destino:
- nome: primeiro nome (ou nome completo se não há coluna separada de sobrenome)
- sobrenome: sobrenome — use null se nome completo estiver em uma única coluna
- cidade: cidade ou município do vereador
- telefone: número de telefone ou celular
- fonte: canal de contato (Ligação, WhatsApp, E-mail etc.) — null se não existir
- status: situação ou resultado do contato — null se não existir
- observacoes: anotações ou comentários — null se não existir

Cabeçalhos disponíveis: {cabecalhos}

Linhas de exemplo:
{json.dumps(amostra, ensure_ascii=False, indent=2)}

Responda apenas com o JSON de mapeamento:"""

    response = client.chat.completions.create(
        model="llama-3.3-70b-versatile",
        messages=[{"role": "user", "content": prompt}],
        temperature=0,
    )

    raw = response.choices[0].message.content.strip()
    raw = raw.replace("```json", "").replace("```", "").strip()
    return json.loads(raw)


def importar_planilha(path_antigo, path_novo, api_key, callback_progresso=None):
    wb = openpyxl.load_workbook(path_antigo, read_only=True, data_only=True)

    cabecalhos = None
    todas_linhas = []

    for nome_aba in wb.sheetnames:
        ws = wb[nome_aba]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue
        if cabecalhos is None:
            cabecalhos = [
                str(c).strip() if c is not None else f"Coluna{i}"
                for i, c in enumerate(rows[0])
            ]
            todas_linhas.extend(rows[1:])
        else:
            todas_linhas.extend(rows[1:])

    wb.close()

    if not cabecalhos:
        raise ValueError("A planilha não contém dados ou está vazia.")

    todas_linhas = [
        r for r in todas_linhas
        if any(c is not None and str(c).strip() for c in r)
    ]

    if not todas_linhas:
        raise ValueError("A planilha não contém linhas de dados.")

    amostra = []
    for row in todas_linhas[:5]:
        d = {cabecalhos[i]: str(row[i]).strip() if i < len(row) and row[i] is not None else ""
             for i in range(len(cabecalhos))}
        amostra.append(d)

    if callback_progresso:
        callback_progresso(f"Analisando estrutura com IA ({len(todas_linhas)} registros encontrados)...")

    mapeamento = _mapear_colunas(cabecalhos, amostra, api_key)

    garantir_excel(path_novo)
    importados = 0
    erros = []

    for i, row in enumerate(todas_linhas, start=2):
        try:
            row_dict = {
                cabecalhos[j]: str(row[j]).strip() if j < len(row) and row[j] is not None else ""
                for j in range(len(cabecalhos))
            }

            def get(campo):
                col = mapeamento.get(campo)
                return row_dict.get(col, "") if col and col in row_dict else ""

            nome_raw = get("nome")
            if not nome_raw:
                continue

            if mapeamento.get("sobrenome"):
                nome_final = nome_raw
                sobrenome_final = get("sobrenome")
            else:
                partes = nome_raw.split(" ", 1)
                nome_final = partes[0]
                sobrenome_final = partes[1] if len(partes) > 1 else ""

            dados = {
                "nome":        nome_final,
                "sobrenome":   sobrenome_final,
                "cidade":      get("cidade"),
                "telefone":    get("telefone"),
                "fonte":       get("fonte"),
                "status":      get("status"),
                "observacoes": get("observacoes"),
            }
            dados["aba"] = _aba_para_status(dados["status"])

            adicionar_linha(path_novo, dados)
            importados += 1

            if callback_progresso and importados % 10 == 0:
                callback_progresso(f"Importando... {importados} de {len(todas_linhas)}")

        except Exception as exc:
            erros.append(f"Linha {i}: {exc}")

    return importados, erros
