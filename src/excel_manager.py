import os
from datetime import datetime
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

ABAS = ["Falhas e Sem Contato", "Sem Interesse"]
CABECALHOS = ["Nome", "Sobrenome", "Cidade", "Telefone", "Fonte", "Status", "Observações", "Data/Hora"]
LARGURAS =   [16,     16,         16,       15,        14,      22,       42,              18]

COR_CABECALHO  = "1C1C3A"
COR_LINHA_PAR  = "E8EDFB"
COR_LINHA_IMPAR = "FFFFFF"


def garantir_excel(path: str) -> None:
    """Cria o arquivo e as abas se ainda não existirem."""
    try:
        wb = openpyxl.load_workbook(path)
        criou_aba = False
        for aba in ABAS:
            if aba not in wb.sheetnames:
                _criar_aba(wb, aba)
                criou_aba = True
        if criou_aba:
            _salvar_com_erro_claro(wb, path)
        return
    except FileNotFoundError:
        pass

    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    for aba in ABAS:
        _criar_aba(wb, aba)
    _salvar_com_erro_claro(wb, path)


def adicionar_linha(path: str, dados: dict) -> None:
    """Insere uma linha na aba correta com data/hora automática."""
    garantir_excel(path)

    wb = openpyxl.load_workbook(path)
    nome_aba = dados.get("aba", ABAS[0])
    if nome_aba not in wb.sheetnames:
        nome_aba = ABAS[0]
    ws = wb[nome_aba]

    linha = [
        dados.get("nome", ""),
        dados.get("sobrenome", ""),
        dados.get("cidade", ""),
        dados.get("telefone", ""),
        dados.get("fonte", ""),
        dados.get("status", ""),
        dados.get("observacoes", ""),
        datetime.now().strftime("%d/%m/%Y %H:%M"),
    ]
    ws.append(linha)

    num = ws.max_row
    cor = COR_LINHA_PAR if num % 2 == 0 else COR_LINHA_IMPAR
    fill = PatternFill(fill_type="solid", fgColor=cor)
    align = Alignment(vertical="center", wrap_text=True)

    for cell in ws[num]:
        cell.fill = fill
        cell.alignment = align

    ws.row_dimensions[num].height = 18
    _salvar_com_erro_claro(wb, path)


def contar_registros_hoje(path: str) -> int:
    """Conta linhas de dados registradas hoje em todas as abas."""
    if not os.path.exists(path):
        return 0
    hoje = datetime.now().strftime("%d/%m/%Y")
    total = 0
    try:
        wb = openpyxl.load_workbook(path, read_only=True)
        for aba in ABAS:
            if aba in wb.sheetnames:
                for row in wb[aba].iter_rows(min_row=2, values_only=True):
                    data_hora = row[7] if len(row) > 7 else None
                    if data_hora and str(data_hora).startswith(hoje):
                        total += 1
        wb.close()
    except Exception:
        pass
    return total


# ── helpers ──────────────────────────────────────────────────────────────────

def _criar_aba(wb: openpyxl.Workbook, nome: str) -> None:
    ws = wb.create_sheet(nome)

    font_cab   = Font(bold=True, color="FFFFFF", name="Calibri", size=10)
    fill_cab   = PatternFill(fill_type="solid", fgColor=COR_CABECALHO)
    align_cab  = Alignment(horizontal="center", vertical="center")

    ws.append(CABECALHOS)
    for cell in ws[1]:
        cell.font      = font_cab
        cell.fill      = fill_cab
        cell.alignment = align_cab

    for col, (cell, largura) in enumerate(zip(ws[1], LARGURAS), start=1):
        ws.column_dimensions[cell.column_letter].width = largura

    ws.row_dimensions[1].height = 22
    ws.freeze_panes = "A2"


def _salvar_com_erro_claro(wb: openpyxl.Workbook, path: str) -> None:
    try:
        wb.save(path)
    except PermissionError:
        nome = Path(path).name
        raise PermissionError(
            f"O arquivo '{nome}' está aberto em outro programa.\n"
            "Feche o Excel e tente salvar novamente."
        )
