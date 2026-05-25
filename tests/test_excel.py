import os
import sys
import tempfile
import unittest
from pathlib import Path

import openpyxl

sys.path.insert(0, str(Path(__file__).parent.parent / "src"))
from excel_manager import adicionar_linha, garantir_excel, ABAS, CABECALHOS


def _path_temporario() -> str:
    """Retorna um caminho único que ainda não existe no disco."""
    fd, path = tempfile.mkstemp(suffix=".xlsx")
    os.close(fd)
    os.unlink(path)
    return path


class TestGarantirExcel(unittest.TestCase):

    def setUp(self):
        self.path = _path_temporario()

    def tearDown(self):
        if os.path.exists(self.path):
            os.unlink(self.path)

    def test_cria_arquivo_do_zero(self):
        garantir_excel(self.path)
        self.assertTrue(os.path.exists(self.path))

    def test_cria_ambas_as_abas(self):
        garantir_excel(self.path)
        wb = openpyxl.load_workbook(self.path)
        for aba in ABAS:
            self.assertIn(aba, wb.sheetnames, f'Aba "{aba}" não encontrada')

    def test_cabecalhos_corretos(self):
        garantir_excel(self.path)
        wb = openpyxl.load_workbook(self.path)
        for aba in ABAS:
            valores = [c.value for c in wb[aba][1]]
            self.assertEqual(valores, CABECALHOS,
                             f'Cabeçalhos incorretos na aba "{aba}"')

    def test_idempotente(self):
        """Chamar garantir_excel duas vezes não duplica abas nem apaga dados."""
        garantir_excel(self.path)
        garantir_excel(self.path)
        wb = openpyxl.load_workbook(self.path)
        self.assertEqual(wb.sheetnames, ABAS)


class TestAdicionarLinha(unittest.TestCase):

    def setUp(self):
        self.path = _path_temporario()

    def tearDown(self):
        if os.path.exists(self.path):
            os.unlink(self.path)

    def _linha(self, aba: str, status: str, nome="Teste", sobrenome="Silva") -> dict:
        return {
            "nome": nome,
            "sobrenome": sobrenome,
            "cidade": "Campinas",
            "telefone": "19999990000",
            "fonte": "Ligação",
            "status": status,
            "observacoes": "Observação de teste",
            "aba": aba,
        }

    # ── Testes de roteamento de aba ──────────────────────────────────────────

    def test_nao_atendida_vai_para_falhas(self):
        dados = self._linha("Falhas e Sem Contato", "Não atendida")
        adicionar_linha(self.path, dados)

        wb = openpyxl.load_workbook(self.path)
        ws_falhas = wb["Falhas e Sem Contato"]
        ws_si     = wb["Sem Interesse"]

        self.assertEqual(ws_falhas.max_row, 2, "Deve ter 1 linha de dados em Falhas")
        self.assertEqual(ws_si.max_row, 1,     "Sem Interesse deve estar vazia")

    def test_sem_interesse_vai_para_aba_correta(self):
        dados = self._linha("Sem Interesse", "Sem interesse")
        adicionar_linha(self.path, dados)

        wb = openpyxl.load_workbook(self.path)
        ws_falhas = wb["Falhas e Sem Contato"]
        ws_si     = wb["Sem Interesse"]

        self.assertEqual(ws_si.max_row, 2,     "Deve ter 1 linha de dados em Sem Interesse")
        self.assertEqual(ws_falhas.max_row, 1, "Falhas deve estar vazia")

    # ── Testes de conteúdo da linha ──────────────────────────────────────────

    def test_campos_salvos_corretamente(self):
        dados = self._linha("Falhas e Sem Contato", "Não atendida",
                            nome="João", sobrenome="Pereira")
        adicionar_linha(self.path, dados)

        ws  = openpyxl.load_workbook(self.path)["Falhas e Sem Contato"]
        row = [c.value for c in ws[2]]

        self.assertEqual(row[0], "João")           # Nome
        self.assertEqual(row[1], "Pereira")        # Sobrenome
        self.assertEqual(row[2], "Campinas")       # Cidade
        self.assertEqual(row[3], "19999990000")    # Telefone
        self.assertEqual(row[4], "Ligação")        # Fonte
        self.assertEqual(row[5], "Não atendida")   # Status
        self.assertEqual(row[6], "Observação de teste")  # Observações
        self.assertIsNotNone(row[7])               # Data/Hora preenchida

    def test_multiplas_linhas_acumulam(self):
        for i in range(3):
            adicionar_linha(self.path,
                            self._linha("Falhas e Sem Contato", "Ocupado",
                                        nome=f"Vereador{i}"))

        ws = openpyxl.load_workbook(self.path)["Falhas e Sem Contato"]
        self.assertEqual(ws.max_row, 4)  # 1 cabeçalho + 3 dados

    def test_aba_invalida_cai_em_falhas(self):
        dados = self._linha("Aba Inexistente", "Qualquer status")
        adicionar_linha(self.path, dados)

        ws = openpyxl.load_workbook(self.path)["Falhas e Sem Contato"]
        self.assertEqual(ws.max_row, 2)


if __name__ == "__main__":
    unittest.main(verbosity=2)
